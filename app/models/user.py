"""
In-memory user model and store for authentication and authorization.
"""
from datetime import datetime
from typing import Dict, Optional, List
from werkzeug.security import generate_password_hash, check_password_hash
from flask_login import UserMixin
import os

try:
    from openpyxl import Workbook, load_workbook
except Exception:  # openpyxl may not be installed yet
    Workbook = None
    load_workbook = None


class User(UserMixin):
    """User account stored in memory (no database)."""

    def __init__(self, username: str, email: str, password: Optional[str] = None, is_admin: bool = False):
        self.username = username
        self.email = email
        self.password_hash = None
        if password:
            self.set_password(password)
        self.created_at = datetime.utcnow()
        self.is_admin = is_admin

    def set_password(self, password: str) -> None:
        self.password_hash = generate_password_hash(password)

    def check_password(self, password: str) -> bool:
        if not self.password_hash:
            return False
        return check_password_hash(self.password_hash, password)

    def get_id(self) -> str:
        # Use username as the unique identifier for Flask-Login
        return self.username

    def __repr__(self) -> str:
        return f'<User {self.username}>'


class UserStore:
    """Simple in-memory store for users keyed by username and email."""

    _users_by_username: Dict[str, User] = {}
    _users_by_email: Dict[str, User] = {}

    @classmethod
    def add(cls, user: User) -> User:
        cls._users_by_username[user.username] = user
        cls._users_by_email[user.email] = user
        return user

    @classmethod
    def get_by_username(cls, username: str) -> Optional[User]:
        return cls._users_by_username.get(username)

    @classmethod
    def get_by_email(cls, email: str) -> Optional[User]:
        return cls._users_by_email.get(email)

    @classmethod
    def exists_username(cls, username: str) -> bool:
        return username in cls._users_by_username

    @classmethod
    def exists_email(cls, email: str) -> bool:
        return email in cls._users_by_email


class ExcelUserStore:
    """Excel-backed user store using openpyxl.

    File schema (first row as headers):
    username | email | password_hash | created_at_iso | is_admin
    """

    FILE_NAME = os.path.join(os.path.dirname(__file__), '..', 'users.xlsx')

    @classmethod
    def _file_path(cls) -> str:
        # Resolve normalized absolute path
        return os.path.abspath(cls.FILE_NAME)

    @classmethod
    def _ensure_file(cls) -> None:
        path = cls._file_path()
        os.makedirs(os.path.dirname(path), exist_ok=True)
        if not os.path.exists(path):
            if Workbook is None:
                raise RuntimeError('openpyxl is required for ExcelUserStore')
            wb = Workbook()
            ws = wb.active
            ws.title = 'users'
            ws.append(['username', 'email', 'password_hash', 'created_at_iso', 'is_admin'])
            wb.save(path)

    @classmethod
    def _load_rows(cls) -> List[List[str]]:
        cls._ensure_file()
        if load_workbook is None:
            raise RuntimeError('openpyxl is required for ExcelUserStore')
        wb = load_workbook(cls._file_path())
        ws = wb.active
        # Skip header
        rows = []
        first = True
        for row in ws.iter_rows(values_only=True):
            if first:
                first = False
                continue
            rows.append(list(row))
        return rows

    @classmethod
    def _append_row(cls, row: List[str]) -> None:
        cls._ensure_file()
        wb = load_workbook(cls._file_path())
        ws = wb.active
        ws.append(row)
        wb.save(cls._file_path())

    @classmethod
    def update_user(cls, original_username: str, updated: User) -> bool:
        """Update an existing user identified by original_username.

        Returns True if a row was updated.
        """
        cls._ensure_file()
        wb = load_workbook(cls._file_path())
        ws = wb.active
        found = False
        # Iterate including header; skip header by index
        for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            username_cell = ws.cell(row=idx, column=1)
            if username_cell.value == original_username:
                ws.cell(row=idx, column=1, value=updated.username)
                ws.cell(row=idx, column=2, value=updated.email)
                ws.cell(row=idx, column=3, value=updated.password_hash or '')
                # keep created_at as-is (column 4)
                ws.cell(row=idx, column=5, value='1' if updated.is_admin else '0')
                found = True
                break
        if found:
            wb.save(cls._file_path())
        return found

    @classmethod
    def add(cls, user: User) -> User:
        # Prevent duplicates
        if cls.exists_username(user.username) or cls.exists_email(user.email):
            return user
        cls._append_row([
            user.username,
            user.email,
            user.password_hash or '',
            user.created_at.isoformat(),
            '1' if user.is_admin else '0',
        ])
        return user

    @classmethod
    def get_by_username(cls, username: str) -> Optional[User]:
        for r in cls._load_rows():
            if r and r[0] == username:
                u = User(username=r[0], email=r[1])
                u.password_hash = r[2] or None
                u.is_admin = (str(r[4]) == '1')
                # created_at is informational; not restoring exact dt
                return u
        return None

    @classmethod
    def get_by_email(cls, email: str) -> Optional[User]:
        for r in cls._load_rows():
            if r and r[1] == email:
                u = User(username=r[0], email=r[1])
                u.password_hash = r[2] or None
                u.is_admin = (str(r[4]) == '1')
                return u
        return None

    @classmethod
    def exists_username(cls, username: str) -> bool:
        return cls.get_by_username(username) is not None

    @classmethod
    def exists_email(cls, email: str) -> bool:
        return cls.get_by_email(email) is not None

